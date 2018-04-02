from openpyxl import load_workbook
# This import used to convert to correct number of significant figures
import to_precision

SIG_FIGS_HDR = 'SigFigs'
RESULTS_HDR = 'GSI_Result'

class Book(object):
    def __init__(self, path, sheet):
        # Create workbook and sheet objects
        self.wb = load_workbook(filename = path)
        self.ws = self.wb[sheet]
        # Create instance variable to store location of relevant columns
        self.col_dict = {}

    def get_col_num(self, headers):
        """Given a header, creates dictionary {k=header, v=column number}
        :type header: List[str]
        :rtype: None
        """
        # Create dict of header string and column number, store as instance var
        col_indices = {cell.value:n for n, cell in enumerate(list(self.ws.rows)[0]) if cell.value in headers}
        self.col_dict = col_indices

    def sig_figify(self, row):
        """Given a row object, uses class instance dict to modify sig figs for
        that row
        :type row: <class 'generator'>
        :rtype: tuple(str)
        """
        # Grab the sig figs for a given row
        sig_figs = int(row[self.col_dict['SigFigs']].value)
        # Grab current value for given row
        curr_val = float(row[self.col_dict['GSI_Result']].value)
        # Modify current value to meet sig fig requirements
        sig_figged = to_precision.std_notation(curr_val, sig_figs)
        # Return tuple of current value, the sig figs, and the new val for the row
        return (curr_val, sig_figs, sig_figged)

    def write_to_sheet(self, results_dict, fpath, outsheet):
        """Writes results dict into an output sheet
        :type results_dict: dict
        :type fpath: str
        :type outsheet: <class 'generator'>
        :rtype: None
        """
        # Make list of dict keys
        dict_keys = [key for key in results_dict.keys()]
        # Write dict keys into sheet
        for i, val in enumerate(dict_keys):
            outsheet.cell(row=1, column=i+1).value = val
        # Write all values into sheet
        for row in range(len(results_dict['Old_val'])):
            outsheet.cell(row=row+1, column=1).value = results_dict['Old_val'][row]
            outsheet.cell(row=row+1, column=2).value = results_dict['Sig_figs'][row]
            outsheet.cell(row=row+1, column=3).value = results_dict['New_val'][row]
        self.wb.save(fpath)

def main():
    """Modifies an Excel worksheet to adhere to sigfigs
    :rtype: None
    """
    # Get workbook filepath and sheet name from user
    fpath = input("Full path to the workbook: ")
    inputs_sheet = input("Name of the worksheet to search and modify: ")
    book = Book(fpath, inputs_sheet)
    book.wb.create_sheet('tmp_sigfig_out')
    # Create a destination sheet for results
    out_sheet = book.wb['tmp_sigfig_out']
    # Make dict for storing old values, sig figs, new values
    output_dict = {'Old_val': [], 'Sig_figs': [], 'New_val': []}
    # Get column numbers for sig figs specifying column and result value column
    book.get_col_num([SIG_FIGS_HDR, RESULTS_HDR])
    # For rows after header
    for row in book.ws.iter_rows(min_row=2):
        # Calculate new value to correct sig figs and store in output_dict
        old, sigs, new = book.sig_figify(row)
        output_dict['Old_val'].append(old)
        output_dict['Sig_figs'].append(sigs)
        output_dict['New_val'].append(new)
    # Write results into destination sheet
    book.write_to_sheet(output_dict, fpath, out_sheet)
    book.wb.close()
    print('Write operation successful!')


if __name__ == '__main__':
    main()
